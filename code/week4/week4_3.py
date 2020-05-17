# openpyxl을 가져옵니다.
import openpyxl as xl

wb = xl.Workbook()

# sheet1은 활성화되어 있는 시트를 선택
sheet1 = wb.active
# sheet1의 시트이름 변경
sheet1.title = "1st sheet"

# 새로운 시트 만들고 sheet2에 저장
sheet2 = wb.create_sheet("2nd sheet")

# sheet1과 sheet2에 동시에 데이터 쓰기
for i in range(1, 10):
    sheet1.cell(row=i, column=1).value = i
    sheet2.cell(row=1, column=i).value = i

wb.save('test.xlsx')